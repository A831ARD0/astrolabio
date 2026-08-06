"""
Exportacion a Excel y CSV.

Lo que se protege: que el archivo no se salte la seguridad por fila, que diga de
donde salio, y que los numeros lleguen como numeros y no como texto.
"""

import csv
import io

from openpyxl import load_workbook


def _abrir(contenido: bytes):
    return load_workbook(io.BytesIO(contenido))


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #

def test_exporta_a_excel_con_etiquetas_del_modelo(cliente, cab_admin, modelo_id):
    r = cliente.post(f"/api/modelos/{modelo_id}/exportar", headers=cab_admin, json={
        "dimensiones": ["cat_sucursal.sucursal_nombre"],
        "metricas": ["monto_venta", "unidades_vendidas"],
        "titulo": "Venta por sucursal", "limite": 100,
    })
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]
    assert "Venta por sucursal" in r.headers["content-disposition"]

    libro = _abrir(r.content)
    hoja = libro["Venta por sucursal"]
    # Encabezados legibles, no nombres tecnicos: quien recibe el archivo no tiene
    # por que traducir 'cat_sucursal.sucursal_nombre'.
    assert [c.value for c in hoja[1]] == ["Sucursal", "Venta", "Unidades"]
    assert hoja.max_row > 1
    assert hoja.freeze_panes == "A2"


def test_los_numeros_llegan_como_numeros(cliente, cab_admin, modelo_id):
    """
    Un Excel con las cifras como texto no se puede sumar ni ordenar, y es el
    motivo numero uno por el que alguien vuelve a teclear los datos a mano.
    """
    r = cliente.post(f"/api/modelos/{modelo_id}/exportar", headers=cab_admin, json={
        "dimensiones": ["cat_sucursal.sucursal_nombre"],
        "metricas": ["monto_venta"], "titulo": "Numeros", "limite": 10,
    })
    hoja = _abrir(r.content)["Numeros"]
    valor = hoja.cell(row=2, column=2).value
    assert isinstance(valor, (int, float)), f"llego como {type(valor)}"
    assert "#,##0" in hoja.cell(row=2, column=2).number_format


def test_el_archivo_dice_de_donde_salio(cliente, cab_admin, modelo_id):
    """
    Un numero en un Excel sin procedencia es un numero que alguien va a discutir
    en una junta sin poder comprobarlo.
    """
    r = cliente.post(f"/api/modelos/{modelo_id}/exportar", headers=cab_admin, json={
        "dimensiones": ["cat_sucursal.sucursal_nombre"],
        "metricas": ["monto_venta"],
        "filtros": [{"campo": "dim_calendario.anio", "op": "=", "valor": 2025}],
        "titulo": "Con procedencia", "limite": 50,
    })
    assert r.status_code == 200, r.text
    hoja = _abrir(r.content)["Procedencia"]
    texto = "\n".join(
        f"{f.value} {g.value}" for f, g in hoja.iter_rows(min_col=1, max_col=2))

    assert "demo_comercial" in texto
    assert "admin@pruebas.example.com" in texto
    assert "monto_venta" in texto
    assert "dim_calendario.anio" in texto, "los filtros aplicados tienen que constar"


def test_avisa_de_las_columnas_personales(cliente, cab_admin, modelo_id):
    """
    Exportar es la via natural para que un dato personal se vaya de la
    herramienta. El archivo lo dice y la auditoria lo registra.
    """
    r = cliente.post(f"/api/modelos/{modelo_id}/exportar", headers=cab_admin, json={
        "dimensiones": ["dim_cliente.cliente_nombre"],
        "metricas": ["monto_venta"], "titulo": "Clientes", "limite": 20,
    })
    assert r.status_code == 200, r.text
    hoja = _abrir(r.content)["Procedencia"]
    texto = "\n".join(str(f.value) for f, in hoja.iter_rows(min_col=1, max_col=1))
    assert "datos personales" in texto.lower()

    from sqlalchemy import select

    from app.db import CrearSesion
    from app.modelos_db import Auditoria

    with CrearSesion() as s:
        fila = s.scalars(
            select(Auditoria).where(Auditoria.accion == "exportacion")
            .order_by(Auditoria.id.desc())
        ).first()
    assert fila is not None
    assert "dim_cliente.cliente_nombre" in fila.detalle["columnas_personales"]


# --------------------------------------------------------------------------- #
# CSV
# --------------------------------------------------------------------------- #

def test_el_csv_lleva_bom_para_que_excel_lea_los_acentos(cliente, cab_admin,
                                                         modelo_id):
    r = cliente.post(f"/api/modelos/{modelo_id}/exportar", headers=cab_admin, json={
        "dimensiones": ["cat_sucursal.sucursal_nombre"],
        "metricas": ["monto_venta"], "formato": "csv", "titulo": "Sucursales",
        "limite": 10,
    })
    assert r.status_code == 200, r.text
    assert r.content.startswith(b"\xef\xbb\xbf"), "sin BOM Excel rompe los acentos"
    assert r.headers["content-type"].startswith("text/csv")

    texto = r.content.decode("utf-8-sig")
    filas = list(csv.reader(io.StringIO(texto)))
    assert filas[0] == ["Sucursal", "Venta"]
    assert len(filas) > 1


# --------------------------------------------------------------------------- #
# Seguridad
# --------------------------------------------------------------------------- #

def test_la_exportacion_respeta_la_seguridad_por_fila(cliente, cab_lector,
                                                      cab_admin, modelo_id):
    """
    Si el archivo se saltara las politicas, exportar seria la forma mas facil de
    sacar lo que no te toca.
    """
    cuerpo = {"dimensiones": ["cat_sucursal.sucursal_nombre"],
              "metricas": ["monto_venta"], "titulo": "Prueba", "limite": 500}

    del_admin = cliente.post(f"/api/modelos/{modelo_id}/exportar",
                             headers=cab_admin, json=cuerpo)
    del_lector = cliente.post(f"/api/modelos/{modelo_id}/exportar",
                              headers=cab_lector, json=cuerpo)
    assert del_lector.status_code == 200, del_lector.text

    filas_admin = _abrir(del_admin.content)["Prueba"].max_row
    hoja_lector = _abrir(del_lector.content)["Prueba"]
    assert hoja_lector.max_row == 2, "el lector regional solo tiene una sucursal"
    assert filas_admin > hoja_lector.max_row

    # Y el archivo del lector lo dice, para que dos numeros distintos no acaben
    # en una discusion: son dos vistas legitimas.
    texto = "\n".join(
        f"{f.value} {g.value}"
        for f, g in _abrir(del_lector.content)["Procedencia"]
        .iter_rows(min_col=1, max_col=2))
    assert "Seguridad por fila" in texto
    assert "rls_por_region" in texto


def test_un_filtro_ambiguo_no_produce_archivo(cliente, cab_admin, modelo_id):
    """Mejor sin archivo que con un archivo de cifras equivocadas."""
    r = cliente.post(f"/api/modelos/{modelo_id}/exportar", headers=cab_admin, json={
        "dimensiones": ["cat_marca.marca_nombre"], "metricas": ["monto_venta"],
        "titulo": "Ambiguo",
    })
    assert r.status_code == 422
    assert len(r.json()["detail"]["rutas"]) == 2


def test_el_tope_de_filas_se_respeta(cliente, cab_admin, modelo_id):
    r = cliente.post(f"/api/modelos/{modelo_id}/exportar", headers=cab_admin, json={
        "metricas": ["monto_venta"], "limite": 999_999,
    })
    assert r.status_code == 422


def test_el_nombre_del_archivo_no_admite_rutas(cliente, cab_admin, modelo_id):
    """Un titulo con barras no debe poder escribir una cabecera rara."""
    r = cliente.post(f"/api/modelos/{modelo_id}/exportar", headers=cab_admin, json={
        "metricas": ["monto_venta"], "titulo": '../../etc/passwd"; x',
        "limite": 10,
    })
    assert r.status_code == 200
    disp = r.headers["content-disposition"]
    assert ".." not in disp and "/" not in disp
