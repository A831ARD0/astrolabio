"""
Exportación a Excel y CSV.

Se genera en el servidor, no en el navegador, por tres razones:

1. El navegador solo tiene las filas que el widget pidió para dibujarse (su
   `limite`). Exportar desde ahí daría un archivo recortado con toda la pinta de
   estar completo.
2. Tiene que pasar por `ejecutar_consulta`, que es el único camino que aplica la
   seguridad por fila. Un exportador propio sería la forma más fácil de sacar
   datos que no te tocan.
3. Los formatos y las etiquetas viven en el modelo. Un Excel con `1234567.891` en
   vez de `$1,234,567.89` obliga a formatear a mano cada vez.

Cada archivo lleva una **hoja de procedencia** con el modelo, su versión, los
filtros aplicados, quién exportó y cuándo. Un número en un Excel sin procedencia es
un número que alguien va a discutir en una junta sin poder comprobarlo.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from semantic.engine import Modelo

# Formatos de número de Excel por formato del modelo.
FORMATOS_EXCEL = {
    "moneda": '"$"#,##0.00',
    "entero": "#,##0",
    "porcentaje": "0.0%",
    "numero": "#,##0.00",
}

# Tope de filas. No es una limitación técnica de openpyxl (aguanta un millón):
# es que un Excel de 200,000 filas ya no se usa para leer, se usa para volver a
# procesar, y para eso hay mejores caminos que un adjunto de correo.
TOPE_FILAS = 200_000


@dataclass
class Procedencia:
    """De dónde salió este archivo. Se escribe en su propia hoja."""
    modelo: str
    version: int
    dimensiones: list[str]
    metricas: list[str]
    filtros: list[dict]
    rutas_elegidas: dict[str, str]
    politicas_aplicadas: list[str]
    email_usuario: str
    filas: int
    sql: str | None = None


def _etiquetas_y_formatos(
    modelo: Modelo, columnas: list[str]
) -> tuple[list[str], list[str | None], list[str]]:
    """
    Para cada columna: su etiqueta legible, su formato de Excel y el nombre de la
    columna PII si lo es.

    Las etiquetas salen del modelo. Exportar `cat_sucursal.sucursal_nombre` como
    encabezado obliga a quien recibe el archivo a traducir nombres técnicos.
    """
    etiquetas: list[str] = []
    formatos: list[str | None] = []
    pii: list[str] = []

    for col in columnas:
        met = modelo.metricas.get(col)
        if met is not None:
            etiquetas.append(met.etiqueta)
            formatos.append(FORMATOS_EXCEL.get(met.formato, FORMATOS_EXCEL["numero"]))
            continue

        if "." in col:
            ent, campo = col.split(".", 1)
            entidad = modelo.entidades.get(ent)
            c = entidad.campos.get(campo) if entidad else None
            if c is not None:
                etiquetas.append(c.etiqueta or c.nombre)
                formatos.append("#,##0" if c.tipo == "entero" else None)
                if c.pii:
                    pii.append(col)
                continue

        etiquetas.append(col)
        formatos.append(None)

    return etiquetas, formatos, pii


# Caracteres que Excel prohíbe en el nombre de una hoja. Sin limpiarlos, un título
# con una barra tumba la exportación entera con un error interno.
_PROHIBIDOS_HOJA = set(r'[]:*?/\\')


def titulo_hoja(titulo: str) -> str:
    limpio = "".join(ch for ch in titulo if ch not in _PROHIBIDOS_HOJA).strip()
    return limpio[:31] or "Datos"


def a_excel(modelo: Modelo, columnas: list[str], filas: list[dict[str, Any]],
            proc: Procedencia, titulo: str) -> tuple[bytes, list[str]]:
    """Devuelve el archivo y las columnas PII que contiene."""
    etiquetas, formatos, pii = _etiquetas_y_formatos(modelo, columnas)

    libro = Workbook()
    hoja = libro.active
    assert hoja is not None
    hoja.title = titulo_hoja(titulo)

    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="374151")
    for i, etiqueta in enumerate(etiquetas, start=1):
        celda = hoja.cell(row=1, column=i, value=etiqueta)
        celda.font = encabezado
        celda.fill = fondo
        celda.alignment = Alignment(horizontal="center")

    for f, fila in enumerate(filas, start=2):
        for i, col in enumerate(columnas, start=1):
            celda = hoja.cell(row=f, column=i, value=fila.get(col))
            if formatos[i - 1] and isinstance(fila.get(col), (int, float)):
                celda.number_format = formatos[i - 1]   # type: ignore[assignment]

    # Congelar el encabezado y poner autofiltro: es lo que cualquiera hace a mano
    # al abrir el archivo.
    hoja.freeze_panes = "A2"
    if filas:
        hoja.auto_filter.ref = (
            f"A1:{get_column_letter(len(columnas))}{len(filas) + 1}")

    for i, etiqueta in enumerate(etiquetas, start=1):
        muestra = [str(fila.get(columnas[i - 1], "")) for fila in filas[:200]]
        largo = max([len(etiqueta)] + [len(m) for m in muestra] or [10])
        hoja.column_dimensions[get_column_letter(i)].width = min(max(largo + 3, 11), 42)

    _hoja_procedencia(libro, proc, pii)
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue(), pii


def _hoja_procedencia(libro: Workbook, proc: Procedencia, pii: list[str]) -> None:
    hoja = libro.create_sheet("Procedencia")
    negrita = Font(bold=True)

    lineas: list[tuple[str, str]] = [
        ("Generado por", "Astrolabio"),
        ("Fecha", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Usuario", proc.email_usuario),
        ("Modelo", f"{proc.modelo} (versión {proc.version})"),
        ("Métricas", ", ".join(proc.metricas) or "—"),
        ("Desglosado por", ", ".join(proc.dimensiones) or "sin desglose (total)"),
        ("Filas", f"{proc.filas:,}".replace(",", ",")),
    ]
    if proc.filtros:
        lineas.append(("Filtros aplicados", ""))
        for f in proc.filtros:
            valor = f.get("valor")
            texto = ", ".join(str(v) for v in valor) if isinstance(valor, list) else str(valor)
            lineas.append((f"    {f.get('campo')}", f"{f.get('op')} {texto}"))
    else:
        lineas.append(("Filtros aplicados", "ninguno"))

    if proc.rutas_elegidas:
        lineas.append(("Caminos elegidos", ""))
        for clave, ruta in proc.rutas_elegidas.items():
            lineas.append((f"    {clave}", ruta))

    if proc.politicas_aplicadas:
        # Que el archivo diga que está recortado por seguridad evita la discusión
        # de "a mí me sale otro número": son dos vistas distintas y legítimas.
        lineas.append(("Seguridad por fila",
                       "SÍ — este archivo solo contiene las filas que el usuario "
                       "puede ver"))
        lineas.append(("Políticas", ", ".join(proc.politicas_aplicadas)))
    if pii:
        lineas.append(("Contiene datos personales", ", ".join(pii)))

    for i, (clave, valor) in enumerate(lineas, start=1):
        c = hoja.cell(row=i, column=1, value=clave)
        c.font = negrita
        hoja.cell(row=i, column=2, value=valor)
    hoja.column_dimensions["A"].width = 26
    hoja.column_dimensions["B"].width = 76


def a_csv(modelo: Modelo, columnas: list[str],
          filas: list[dict[str, Any]]) -> tuple[bytes, list[str]]:
    """
    CSV con BOM de UTF-8 a propósito: sin él, Excel en Windows abre los acentos
    como caracteres raros y la primera reacción es que el archivo está roto.
    """
    etiquetas, _, pii = _etiquetas_y_formatos(modelo, columnas)

    salida = io.StringIO()
    escritor = csv.writer(salida, lineterminator="\r\n")
    escritor.writerow(etiquetas)
    for fila in filas:
        escritor.writerow([fila.get(c) for c in columnas])
    return b"\xef\xbb\xbf" + salida.getvalue().encode("utf-8"), pii


def nombre_archivo(base: str, extension: str) -> str:
    """Nombre seguro para una cabecera HTTP y para un sistema de archivos."""
    limpio = "".join(
        ch if ch.isalnum() or ch in " -_" else "_" for ch in base
    ).strip() or "astrolabio"
    marca = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    return f"{limpio[:60]} {marca}.{extension}"
